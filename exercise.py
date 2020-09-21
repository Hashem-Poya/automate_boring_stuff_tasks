import openpyxl, sys
 

times = int(input('Enter the number to create excel (x, x) columns, rows: '))


wb = openpyxl.Workbook()
sheet = wb['Sheet']

for i in range(1, times + 1):
	sheet.cell(row=i + 1, column=1).value = i
	sheet.cell(row=i + 1, column=1)
	sheet.cell(row=1, column=i+1).value = i
	sheet.cell(row=1, column=i+1)
 
for i in range(2, times + 2):
	for j in range(2, times + 2):
		x = sheet.cell(row=i, column=1).value
		y = sheet.cell(row=1, column=j).value
		sheet.cell(row=i, column=j).value = x * y
 
wb.save('mutliple_table.xlsx')